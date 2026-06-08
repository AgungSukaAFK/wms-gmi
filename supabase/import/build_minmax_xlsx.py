#!/usr/bin/env python3
"""Build a WIDE (one row per part) Excel for editing stock min/max per cabang.

Input  : long CSV exported from VPS with header:
         part_number,part_name,nama_cabang,qty,min_qty,max_qty
Output : .xlsx with columns:
         No. Barang | Deskripsi Barang | <CABANG> QTY | <CABANG> MIN | <CABANG> MAX | ...

QTY columns are read-only reference (greyed). Only MIN/MAX should be edited.
Re-import is done by convert_minmax_xlsx_to_staging.py which parses the
" MIN" / " MAX" column suffixes.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: python3 build_minmax_xlsx.py <input_long_csv> <output_xlsx>", file=sys.stderr)
        return 2

    src = Path(sys.argv[1])
    out = Path(sys.argv[2])
    if not src.exists():
        print(f"Input not found: {src}", file=sys.stderr)
        return 1

    with src.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Collect part order (first seen) and cabang set
    parts: dict[str, dict] = {}
    cabang_order: list[str] = []
    for r in rows:
        pn = (r.get("part_number") or "").strip()
        if not pn:
            continue
        cab = (r.get("nama_cabang") or "").strip()
        if cab and cab not in cabang_order:
            cabang_order.append(cab)
        p = parts.setdefault(pn, {"name": (r.get("part_name") or "").strip(), "cab": {}})
        if not p["name"]:
            p["name"] = (r.get("part_name") or "").strip()

        def num(v):
            v = (v or "").strip()
            if v == "":
                return 0
            try:
                return int(float(v))
            except ValueError:
                return 0

        p["cab"][cab] = (num(r.get("qty")), num(r.get("min_qty")), num(r.get("max_qty")))

    cabang_order.sort()

    wb = Workbook()
    ws = wb.active
    ws.title = "STOCK MIN MAX"

    # ---- header ----
    header = ["No. Barang", "Deskripsi Barang"]
    qty_cols: list[int] = []
    minmax_cols: list[int] = []
    for cab in cabang_order:
        header += [f"{cab} QTY", f"{cab} MIN", f"{cab} MAX"]
    ws.append(header)

    # mark qty vs min/max column indexes (1-based)
    col = 3
    for _ in cabang_order:
        qty_cols.append(col)        # QTY
        minmax_cols.append(col + 1)  # MIN
        minmax_cols.append(col + 2)  # MAX
        col += 3

    grey = PatternFill("solid", fgColor="D9D9D9")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, _ in enumerate(header, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # ---- data ----
    for pn in sorted(parts.keys()):
        p = parts[pn]
        line = [pn, p["name"]]
        for cab in cabang_order:
            q, mn, mx = p["cab"].get(cab, (0, 0, 0))
            line += [q, mn, mx]
        ws.append(line)

    last_row = ws.max_row
    # style columns: QTY grey (reference), MIN/MAX yellow (editable)
    for r in range(2, last_row + 1):
        for c in qty_cols:
            cell = ws.cell(row=r, column=c)
            cell.fill = grey
        for c in minmax_cols:
            cell = ws.cell(row=r, column=c)
            cell.fill = yellow

    # widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    for c in range(3, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # ---- petunjuk sheet ----
    ws2 = wb.create_sheet("PETUNJUK")
    notes = [
        ["PETUNJUK PENGISIAN MIN / MAX STOCK"],
        [""],
        ["1. Kolom abu-abu (QTY) = stok saat ini, HANYA ACUAN. Jangan diubah."],
        ["2. Kolom kuning (MIN / MAX) = silakan diedit sesuai kebutuhan tiap cabang."],
        ["3. JANGAN mengubah / menghapus kolom 'No. Barang' dan 'Deskripsi Barang'."],
        ["4. JANGAN menambah / menghapus / menggeser kolom. Isi angka bulat (>= 0)."],
        ["5. Kosongkan = dianggap 0. Simpan tetap format .xlsx."],
        ["6. Setelah selesai, kirim file ini kembali untuk diimport ke sistem."],
    ]
    for row in notes:
        ws2.append(row)
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 90

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Parts: {len(parts)}  Cabang: {len(cabang_order)} -> {cabang_order}")
    print(f"Output: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
