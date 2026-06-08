#!/usr/bin/env python3
"""Convert the edited WIDE min/max Excel back into a long staging CSV.

Input  : .xlsx produced by build_minmax_xlsx.py and edited by the user.
         Columns: No. Barang | Deskripsi Barang | <CABANG> QTY | <CABANG> MIN | <CABANG> MAX ...
Output : staging CSV with header:
         batch_code,part_number,nama_cabang,min_qty,max_qty,source_row

Only MIN/MAX are imported; QTY columns are ignored.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def to_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s == "":
        return 0
    s = s.replace(".", "").replace(",", ".")  # tolerate Indonesian formatting
    try:
        return int(float(s))
    except ValueError:
        return 0


def main() -> int:
    if len(sys.argv) < 4:
        print(
            "Usage: python3 convert_minmax_xlsx_to_staging.py <input_xlsx> <output_csv> <batch_code>",
            file=sys.stderr,
        )
        return 2

    src = Path(sys.argv[1])
    out = Path(sys.argv[2])
    batch = sys.argv[3]
    if not src.exists():
        print(f"Input not found: {src}", file=sys.stderr)
        return 1

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["STOCK MIN MAX"] if "STOCK MIN MAX" in wb.sheetnames else wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [(str(h).strip() if h is not None else "") for h in header]

    if header[0].lower() not in {"no. barang", "part_number", "part number"}:
        print(f"Unexpected first column: {header[0]!r}", file=sys.stderr)
        return 1

    # Build {cabang: {"MIN": idx, "MAX": idx}}
    cabang_cols: dict[str, dict[str, int]] = {}
    for idx, name in enumerate(header):
        up = name.upper()
        if up.endswith(" MIN"):
            cabang_cols.setdefault(name[:-4].strip(), {})["MIN"] = idx
        elif up.endswith(" MAX"):
            cabang_cols.setdefault(name[:-4].strip(), {})["MAX"] = idx
        # " QTY" columns are intentionally ignored

    cabang = sorted(cabang_cols.keys())
    if not cabang:
        print("No <CABANG> MIN/MAX columns detected in header", file=sys.stderr)
        return 1

    out.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    with out.open("w", encoding="utf-8", newline="") as fo:
        w = csv.writer(fo)
        w.writerow(["batch_code", "part_number", "nama_cabang", "min_qty", "max_qty", "source_row"])
        for rnum, row in enumerate(rows, start=2):
            if not row:
                continue
            pn = (str(row[0]).strip() if row[0] is not None else "")
            if pn == "":
                continue
            for cab in cabang:
                mi = cabang_cols[cab].get("MIN")
                ma = cabang_cols[cab].get("MAX")
                mn = to_int(row[mi]) if mi is not None and mi < len(row) else 0
                mx = to_int(row[ma]) if ma is not None and ma < len(row) else 0
                w.writerow([batch, pn, cab, mn, mx, rnum])
                written += 1

    print(f"Cabang detected: {len(cabang)} -> {cabang}")
    print(f"Rows written: {written}")
    print(f"Output: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
